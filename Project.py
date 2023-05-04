from urllib.request import urlopen, Request
from bs4 import BeautifulSoup

import keys
from twilio.rest import Client

import openpyxl as xl
from openpyxl.styles import Font, Alignment

wb = xl.Workbook()

ws = wb.active

ws.title = 'Crypto Info'

ws['A1'] = 'Currency Name'
ws['B1'] = 'Currency Symbol'
ws['C1'] = 'Current Price'
ws['D1'] = 'Percent Change (24 Hours)'
ws['E1'] = 'Old Price'

myfont = Font(name='Times New Roman', size=24, bold=True, italic=True)

ws['A1'].font = myfont
ws['B1'].font = myfont
ws['C1'].font = myfont
ws['D1'].font = myfont
ws['E1'].font = myfont

write_row = 2
write_colA = 1
write_colB = 2
write_colC = 3
write_colD = 4
write_colE = 5

ws.column_dimensions['A'].width = 27
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 46
ws.column_dimensions['E'].width = 16

url = 'https://www.cryptoslate.com/coins/'

# Request in case 404 Forbidden error
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title.text)

coin = soup.findAll('tr')

for row in coin[1:6]:
    td = row.findAll('td')
    rank = td[0].text
    name_symbol = td[1].text
    words = name_symbol.split()
    name = words[0]
    symbol = words[1]
    current_price = round(float(td[2].text.replace("$","").replace(" ","").replace(",","")),2)
    change = td[3].text.replace(" ","")
    if change[0] == '+':
        new_change = float(td[3].text.replace("%","").replace("+",""))
        percent = round(1+(new_change / 100),4)
    if change[0] == '-':
        new_change = float(td[3].text.replace("%","").replace("-",""))
        percent =  round(1-(new_change / 100),4)
    old_price = round(current_price/percent,2)

    if name == 'Bitcoin':
        if old_price + 5 < current_price:
            message = 'Price for Bitcoin has increased more than $5'
            client = Client(keys.accountSID, keys.auth_token)
            TwilioNumber = '+17207042416'
            mycellphone = '+17203698242'
            textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                                body=message)
        if old_price - 5 > current_price:
            message = 'Price for Bitcoin has decreased more than $5'
            client = Client(keys.accountSID, keys.auth_token)
            TwilioNumber = '+17207042416'
            mycellphone = '+17203698242'
            textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                                body=message)
    elif name == 'Ethereum':
        if old_price + 5 < current_price:
            message = 'Price for Ethereum has increased more than $5'
            client = Client(keys.accountSID, keys.auth_token)
            TwilioNumber = '+17207042416'
            mycellphone = '+17203698242'
            textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                                body=message)
        if old_price - 5 > current_price:
            message = 'Price for Ethereum has decreased more than $5'
            client = Client(keys.accountSID, keys.auth_token)
            TwilioNumber = '+17207042416'
            mycellphone = '+17203698242'
            textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                                    body=message)
            
    ws.cell(write_row, write_colA).value = name
    ws.cell(write_row, write_colB).value = symbol
    ws.cell(write_row, write_colC).value = current_price
    ws.cell(write_row, write_colD).value = change
    ws.cell(write_row, write_colE).value = old_price

    myfont2 = Font(name='Times New Roman', size=16, bold=True, italic=False)
    align = Alignment(horizontal='right')

    ws.cell(write_row, write_colA).font = myfont2
    ws.cell(write_row, write_colC).number_format = u'"$ "#,##0.00'
    ws.cell(write_row, write_colD).number_format = '#,##0.00'
    ws.cell(write_row, write_colD).alignment = align
    ws.cell(write_row, write_colE).number_format = u'"$ "#,##0.00'

    write_row += 1

wb.save('Crypto.xlsx')